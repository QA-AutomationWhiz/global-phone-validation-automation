from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import pandas as pd
import phonenumbers
from phonenumbers import PhoneNumberType
from phonenumbers import example_number_for_type
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import time

# =========================================================
# CONFIG
# =========================================================
URL = "http://172.16.51.206:3000/career-apply/?slug=business-development-executive"

WAIT_TIME = 10

driver = webdriver.Chrome()

driver.maximize_window()

wait = WebDriverWait(driver, WAIT_TIME)

driver.get(URL)

time.sleep(3)

# =========================================================
# START TIMER
# =========================================================
start_time = time.time()

# =========================================================
# LOCATORS
# =========================================================
country_dropdown_xpath = "//select"

contact_input_xpath = "//input[@placeholder='Enter Contact Number']"

apply_button_xpath = "//button[contains(.,'Apply Now')]"

validation_xpath = "//p[contains(text(),'valid')]"

# =========================================================
# GET COUNTRY LIST
# =========================================================
dropdown = wait.until(
    EC.presence_of_element_located(
        (By.XPATH, country_dropdown_xpath)
    )
)

select = Select(dropdown)

countries = select.options

total_countries = len(countries)

print("\n====================================================")
print(f"TOTAL COUNTRIES FOUND : {total_countries}")
print("====================================================")

# =========================================================
# STORE RESULTS
# =========================================================
all_results = []

issue_results = []

# =========================================================
# FUNCTION
# =========================================================
def get_country_code(country_text):

    # Example:
    # IN (+91)

    return country_text.split(" ")[0].strip()

# =========================================================
# START TESTING
# =========================================================
for i in range(total_countries):

    completed = i
    remaining = total_countries - completed

    # =====================================================
    # TIME CALCULATION
    # =====================================================
    current_time = time.time()

    elapsed_time = current_time - start_time

    avg_time_per_country = (
        elapsed_time / completed
        if completed != 0 else 0
    )

    estimated_remaining_seconds = (
        avg_time_per_country * remaining
    )

    estimated_remaining_minutes = round(
        estimated_remaining_seconds / 60,
        2
    )

    try:

        # =================================================
        # REFRESH DROPDOWN
        # =================================================
        dropdown = wait.until(
            EC.presence_of_element_located(
                (By.XPATH, country_dropdown_xpath)
            )
        )

        select = Select(dropdown)

        country_name = select.options[i].text.strip()

        country_code = get_country_code(country_name)

        print("\n====================================================")
        print(f"COUNTRY {i+1}/{total_countries}")
        print(f"TESTING : {country_name}")
        print(f"COMPLETED : {completed}")
        print(f"REMAINING : {remaining}")
        print(f"APPROX REMAINING TIME : {estimated_remaining_minutes} Minutes")
        print("====================================================")

        # =================================================
        # SELECT COUNTRY
        # =================================================
        select.select_by_index(i)

        time.sleep(1)

        # =================================================
        # GET REAL VALID NUMBER
        # =================================================
        valid_number = None

        try:

            example_number = example_number_for_type(
                country_code,
                PhoneNumberType.MOBILE
            )

            if example_number:

                valid_number = str(
                    example_number.national_number
                )

        except Exception:
            pass

        # =================================================
        # NO VALID NUMBER FOUND
        # =================================================
        if not valid_number:

            result = {
                "Country": country_name,
                "Valid Number": "",
                "Test Type": "VALID_NUMBER_NOT_FOUND",
                "Input": "",
                "Validation Message": "No valid sample number found",
                "Issue Found": "YES"
            }

            all_results.append(result)

            issue_results.append(result)

            print(result)

            continue

        # =================================================
        # CREATE INVALID NUMBER
        # =================================================
        invalid_number = "1" * len(valid_number)

        # Agar accidentally valid jesa ban gaya
        if invalid_number == valid_number:
            invalid_number = "9" + invalid_number[1:]

        # =================================================
        # TEST CASES
        # =================================================
        test_cases = [

            # VALID NUMBER
            {
                "test_type": "VALID_NUMBER",
                "value": valid_number,
                "expect_validation": False
            },

            # LESS THAN VALID LENGTH
            {
                "test_type": "LESS_THAN_VALID_LENGTH",
                "value": valid_number[:-1],
                "expect_validation": True
            },

            # INVALID NUMBER
            {
                "test_type": "INVALID_NUMBER",
                "value": invalid_number,
                "expect_validation": True
            }

        ]

        # =================================================
        # RUN TEST CASES
        # =================================================
        for test in test_cases:

            contact_input = wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, contact_input_xpath)
                )
            )

            # Clear input
            contact_input.clear()

            # Enter value
            contact_input.send_keys(test["value"])

            # Click Apply Button
            driver.find_element(
                By.XPATH,
                apply_button_xpath
            ).click()

            time.sleep(1)

            # =================================================
            # CHECK VALIDATION
            # =================================================
            validation_found = False

            validation_message = "No Validation"

            try:

                validation = driver.find_element(
                    By.XPATH,
                    validation_xpath
                )

                if validation.is_displayed():

                    validation_found = True

                    validation_message = validation.text.strip()

            except NoSuchElementException:

                validation_found = False

                validation_message = "No Validation"

            # =================================================
            # ISSUE LOGIC
            # =================================================
            issue_found = False

            # Validation expected but NOT shown
            if test["expect_validation"] and not validation_found:
                issue_found = True

            # Validation NOT expected but shown
            if not test["expect_validation"] and validation_found:
                issue_found = True

            # =================================================
            # SAVE RESULT
            # =================================================
            result = {
                "Country": country_name,
                "Valid Number": valid_number,
                "Test Type": test["test_type"],
                "Input": test["value"],
                "Validation Message": validation_message,
                "Issue Found": "YES" if issue_found else "NO"
            }

            all_results.append(result)

            print(result)

            if issue_found:
                issue_results.append(result)

    except Exception as e:

        error_data = {
            "Country": country_name if 'country_name' in locals() else "UNKNOWN",
            "Valid Number": "",
            "Test Type": "SYSTEM_ERROR",
            "Input": "",
            "Validation Message": str(e),
            "Issue Found": "YES"
        }

        all_results.append(error_data)

        issue_results.append(error_data)

        print(error_data)

# =========================================================
# TOTAL EXECUTION TIME
# =========================================================
end_time = time.time()

total_execution_minutes = round(
    (end_time - start_time) / 60,
    2
)

# =========================================================
# CREATE DATAFRAME
# =========================================================
full_df = pd.DataFrame(all_results)

issue_df = pd.DataFrame(issue_results)

# =========================================================
# FILE NAMES
# =========================================================
full_report_name = "All_Country_Validation_Report.xlsx"

issue_report_name = "Country_Issue_Report.xlsx"

# =========================================================
# EXPORT EXCEL
# =========================================================
full_df.to_excel(full_report_name, index=False)

issue_df.to_excel(issue_report_name, index=False)

# =========================================================
# FORMAT FULL REPORT
# =========================================================
wb = load_workbook(full_report_name)

ws = wb.active

# Header Style
header_fill = PatternFill(
    start_color="FF6B00",
    end_color="FF6B00",
    fill_type="solid"
)

header_font = Font(
    color="FFFFFF",
    bold=True
)

# Pass / Fail Colors
green_fill = PatternFill(
    start_color="CCFFCC",
    end_color="CCFFCC",
    fill_type="solid"
)

red_fill = PatternFill(
    start_color="FFCCCC",
    end_color="FFCCCC",
    fill_type="solid"
)

# Header Style
for cell in ws[1]:

    cell.fill = header_fill

    cell.font = header_font

# Auto Width
for column_cells in ws.columns:

    length = max(
        len(str(cell.value))
        if cell.value else 0
        for cell in column_cells
    )

    column_letter = get_column_letter(
        column_cells[0].column
    )

    ws.column_dimensions[column_letter].width = length + 5

# Highlight Rows
for row in range(2, ws.max_row + 1):

    issue_value = ws[f'F{row}'].value

    if issue_value == "YES":

        for cell in ws[row]:
            cell.fill = red_fill

    else:

        for cell in ws[row]:
            cell.fill = green_fill

# Freeze Header
ws.freeze_panes = "A2"

wb.save(full_report_name)

# =========================================================
# FORMAT ISSUE REPORT
# =========================================================
wb2 = load_workbook(issue_report_name)

ws2 = wb2.active

for cell in ws2[1]:

    cell.fill = header_fill

    cell.font = header_font

for column_cells in ws2.columns:

    length = max(
        len(str(cell.value))
        if cell.value else 0
        for cell in column_cells
    )

    column_letter = get_column_letter(
        column_cells[0].column
    )

    ws2.column_dimensions[column_letter].width = length + 5

for row in range(2, ws2.max_row + 1):

    for cell in ws2[row]:

        cell.fill = red_fill

ws2.freeze_panes = "A2"

wb2.save(issue_report_name)

# =========================================================
# FINAL SUMMARY
# =========================================================
print("\n====================================================")
print("AUTOMATION EXECUTION COMPLETED")
print("====================================================")

print(f"\nTOTAL COUNTRIES : {total_countries}")

print(f"TOTAL RESULTS : {len(full_df)}")

print(f"TOTAL ISSUES : {len(issue_df)}")

print(
    f"SUCCESSFULLY TESTED COUNTRIES : "
    f"{total_countries - len(issue_df['Country'].unique())}"
)

print(
    f"ISSUE COUNTRIES : "
    f"{len(issue_df['Country'].unique())}"
)

print(
    f"\nTOTAL EXECUTION TIME : "
    f"{total_execution_minutes} Minutes"
)

print("\nEXCEL REPORT GENERATED :")

print(f"1. {full_report_name}")

print(f"2. {issue_report_name}")

# =========================================================
# ISSUE COUNTRY LIST
# =========================================================
if len(issue_df) > 0:

    print("\nCOUNTRIES HAVING ISSUES :\n")

    for country in issue_df["Country"].unique():

        print(country)

else:

    print("\nNO ISSUES FOUND")

driver.quit()